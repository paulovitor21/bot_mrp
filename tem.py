import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import numbers
from datetime import datetime
# Caminho do arquivo de origem (.xlsb)
arquivo_origem = r"C:\Users\Paulo\Desktop\tarefa\planilhas_base\Downloads-3\Resultado (4).xlsb"

# Nome da planilha dentro do arquivo
nome_planilha = "Summy Daily_ByLine"

# Abrir o Excel e a planilha
excel = win32.Dispatch("Excel.Application")
excel.Visible = False  # Não mostrar o Excel na tela
workbook = excel.Workbooks.Open(arquivo_origem)

# Selecionar a planilha pelo nome
sheet = workbook.Sheets(nome_planilha)

# Determinar o número de linhas e colunas a serem lidas
ultima_linha = sheet.Cells(sheet.Rows.Count, 1).End(-4162).Row  # Encontrar a última linha
ultima_coluna = sheet.Cells(1, sheet.Columns.Count).End(-4159).Column  # Encontrar a última coluna

print(f"Última linha: {ultima_linha}")
print(f"Última coluna: {ultima_coluna}")

# Ler os dados da planilha
dados = []
for linha in range(1, ultima_linha + 1):
    # Acessa da coluna B até a última coluna
    row_data = sheet.Range(sheet.Cells(linha, 2), sheet.Cells(linha, ultima_coluna)).Value
    
    # Verificando se row_data é None ou se contém dados
    if row_data:
        row_data_flat = [cell for cell in row_data[0]]  # Achata a lista para obter os dados
        dados.append(row_data_flat)
    else:
        print(f"Linha {linha} está vazia ou não contém dados válidos")

# Fechar o Excel
workbook.Close(False)
excel.Quit()

# Verificando se os dados foram lidos corretamente
if not dados:
    print("Nenhum dado foi lido.")
else:
    print(f"Dados lidos: {dados[:5]}")  # Imprime as primeiras 5 linhas lidas

# Manipular os dados (selecionando as colunas B, C e F em diante)
dados_filtrados = []
for row in dados:
    dados_filtrados.append([row[1], row[0]] + row[5:])  # Pega B, C e F até a última coluna

# Remover timezone das datas
for row in dados_filtrados:
    for i, value in enumerate(row):
        if isinstance(value, datetime):
            # Remove o timezone
            row[i] = value.replace(tzinfo=None)

# Criar o arquivo de destino
wb_destino = Workbook()
sheet_destino = wb_destino.active

# Escrever os dados filtrados no novo arquivo
for i, row in enumerate(dados_filtrados):
    for j, value in enumerate(row):
        # Verificação e conversão específica para datas
        if isinstance(value, datetime):
            # Converte para data do Excel
            cell = sheet_destino.cell(row=i+1, column=j+1, value=value)
            # Define o formato de data abreviada do Excel
            cell.number_format = 'DD/MM/YYYY'
        else:
            sheet_destino.cell(row=i+1, column=j+1, value=value)

# Salvar o novo arquivo
wb_destino.save("arquivo_destino.xlsx")

print("Arquivo salvo com sucesso!")