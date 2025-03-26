import pandas as pd
import time
import os
import xlwings as xw
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import numbers
from datetime import datetime
from botcity.core import DesktopBot

class Bot(DesktopBot):
    def action(self, execution=None):
        pass

    def open_excel(self, excel_file_path):
        os.startfile(excel_file_path)
        time.sleep(10)  # Aguardar o Excel carregar completamente

        try:
            wb = xw.Book(excel_file_path)  # Abrir ou pegar o arquivo ativo
            return wb
        except Exception as e:
            print(f"Erro ao abrir o arquivo no Excel: {e}")
            return None

    def process_nfp(self, file_nfp, master_all_path):
        print(f"Processing NFP file: {file_nfp}")

        # Abrir o Excel com win32com para leitura dos dados
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(file_nfp)

        # Nome da planilha dentro do arquivo
        nome_planilha = "Summy Daily_ByLine"

        # Selecionar a planilha pelo nome
        sheet = workbook.Sheets(nome_planilha)

        # Determinar o número de linhas e colunas a serem lidas
        ultima_linha = sheet.Cells(sheet.Rows.Count, 1).End(-4162).Row
        ultima_coluna = sheet.Cells(1, sheet.Columns.Count).End(-4159).Column

        print(f"Última linha: {ultima_linha}")
        print(f"Última coluna: {ultima_coluna}")

        # Ler os dados da planilha
        dados = []
        for linha in range(1, ultima_linha + 1):
            row_data = sheet.Range(sheet.Cells(linha, 2), sheet.Cells(linha, ultima_coluna)).Value
            
            if row_data:
                row_data_flat = [cell for cell in row_data[0]]
                dados.append(row_data_flat)
            else:
                print(f"Linha {linha} está vazia ou não contém dados válidos")

        # Fechar o Excel
        workbook.Close(False)
        excel.Quit()

        # Manipular os dados (selecionando as colunas B, C e F em diante)
        dados_filtrados = []
        for row in dados:
            dados_filtrados.append([row[1], row[0]] + row[5:])  # Pega B, C e F até a última coluna

        # Remover timezone das datas
        for row in dados_filtrados:
            for i, value in enumerate(row):
                if isinstance(value, datetime):
                    row[i] = value.replace(tzinfo=None)

        # Criar o arquivo de destino temporário
        wb_destino = Workbook()
        sheet_destino = wb_destino.active

        # Escrever os dados filtrados no novo arquivo
        for i, row in enumerate(dados_filtrados):
            for j, value in enumerate(row):
                if isinstance(value, datetime):
                    cell = sheet_destino.cell(row=i+1, column=j+1, value=value)
                    cell.number_format = 'DD/MM/YYYY'
                else:
                    sheet_destino.cell(row=i+1, column=j+1, value=value)

        # Salvar o arquivo temporário
        wb_destino.save("temp_nfp.xlsx")

        # Abrir o Excel e carregar o workbook existente
        wb = self.open_excel(master_all_path)
        if not wb:
            return

        ws = wb.sheets["PlanPPH"]  # Selecionar a aba desejada
        ws.activate()
        print(f"Aba ativa: {ws.name}")

        # Selecionar e limpar os dados antigos, mantendo o cabeçalho
        ws.range("A1").expand().clear_contents()

        # Carregar os dados do arquivo temporário
        nfp_df = pd.read_excel("temp_nfp.xlsx", header=0)
        # Renomear colunas 'Unnamed'
        nfp_df.columns = ['' if 'Unnamed:' in str(col) else col.split('.')[0] for col in nfp_df.columns]

        # nfp_df.columns = ['' if 'Unnamed:' in str(col) else col for col in nfp_df.columns]
        # Copiar os dados para a área de transferência
        nfp_df.to_clipboard(index=False, header=True)
        

        # Colar os dados no Excel
        ws.range("A1").api.PasteSpecial(Paste=-4104)  # xlPasteAll

        # Salvar e fechar o arquivo
        try:
            wb.save()
            wb.close()
            print("NFP data has been updated in master_all.xlsx")
            
            # Remover o arquivo temporário
            os.remove("temp_nfp.xlsx")
        except Exception as e:
            print(f"Erro ao salvar/fechar o arquivo: {e}")

if __name__ == '__main__':
    bot = Bot()
    file_nfp = r"C:\Users\Paulo\Desktop\tarefa\planilhas_base\NFP 03062025 0935.xlsx"
    master_all_path = r"C:\Users\Paulo\Desktop\bot_mrp\bot_mrp\05.03_Master_All_Sourcing_.xlsb"
    bot.process_nfp(file_nfp, master_all_path)